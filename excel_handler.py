
from openpyxl import Workbook, load_workbook

def escribir_excel(nombre_archivo, datos):
    """
    Escribe los datos en un archivo Excel.
    :param nombre_archivo: Nombre del archivo Excel.
    :param datos: Datos a escribir.
    """
    libro = Workbook()
    hoja = libro.active


    encabezados = list(datos.keys())
    hoja.append(encabezados)


    valores = list(datos.values())
    hoja.append(valores)

    # Guardar el libro
    libro.save(nombre_archivo)

def leer_excel(nombre_archivo):
    """
    Lee los datos desde un archivo Excel.
    :param nombre_archivo: Nombre del archivo Excel.
    :return: Datos leídos desde el archivo Excel.
    """
    libro = load_workbook(nombre_archivo)
    hoja = libro.active

    encabezados = [celda.value for celda in hoja[1]]

    datos = {}
    for columna, encabezado in zip(hoja.iter_cols(), encabezados):
        valores = [celda.value for celda in columna[1:]]
        datos[encabezado] = valores

    return datos
